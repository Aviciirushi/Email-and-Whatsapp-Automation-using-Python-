import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import os
import re
import random
import time

# -----------------------
# 1. Configuration
# -----------------------
EMAIL_ADDRESS = "youremial@email.com"
EMAIL_PASSWORD = "App_Password"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EXCEL_FILE = "sent_emails_list.xlsx"
MAX_EMAILS_PER_DAY = 100  # Gmail safe daily send limit

# -----------------------
# 2. Email Templates
# -----------------------
def get_email_template(day):
    subject, body = "", ""

    if day == 1:
        subject = "Your quotation from Digimed Exxim"
        body = """Dear Sir/Mam,

I hope you are doing well. I just wanted to follow up on the quotation we recently shared.

At *Digimed Exxim*, we specialize in exporting:
✅ Steroids & Weight Loss Products
✅ Tapentadol & Carisoprodol
✅ Exclusive Indian Ayurvedic & Herbal Formulations (our USP)

Why our partners choose us:
* WHO-GMP Certified Products
* Shelf Life of 15+ months
* Flexible MOQs tailored to your market needs

👉 Would you like me to prepare a revised proposal that best fits your requirements?

Warm regards,
Rahul Singh
Founder, Digimed Exxim
🌐 www.digimedexxim.com
📧 digimedexxim@gmail.com
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    elif day == 4:
        subject = "How distributors in USA & UK scaled with Digimed Exxim"
        body = """Dear Sir/Mam,

Just checking in again. Many of our clients in the *USA, UK, and Australia* rely on us for a consistent supply of *Weight Loss products, Steroids, and Tapentadol*, while markets like *France, Singapore, and Malaysia* prefer our *Ayurvedic & Herbal range*.

📌 Example: A distributor in USA reduced delays by *30% in 3 months* after switching to our supply chain.

👉 Would you like me to share a *region-specific case study* to show how we can add the same value to your business?

Best regards,
Rahul Singh
Digimed Exxim
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    elif day == 10:
        subject = "Why our herbal + pharma mix gives distributors an edge"
        body = """Dear Sir/Mam,

What sets Digimed Exxim apart isn’t just our strong pharma portfolio (Steroids, Weight Loss, Tapentadol, Carisoprodol) — but also our *exclusive Indian Herbal & Ayurvedic line*, which has been a game-changer for distributors.

Benefits for you:
✅ Strong margins + fast-growing natural demand
✅ WHO-GMP certified & 15+ months shelf life
✅ Timely shipments across USA, UK, Australia, France, Singapore & more
✅ Flexible MOQ for faster market entry

👉 Would you like me to prepare a *supply plan tailored to your market’s demand cycles*?

Warm regards,
Rahul Singh
Digimed Exxim
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    elif day == 20:
        subject = "High demand for Weight Loss & Herbal products this quarter"
        body = """Dear Sir/Mam,

We’re seeing a strong surge in demand for *Weight Loss Products, Tapentadol, and Herbal formulations* in markets like *USA, France, and Australia*.

To help distributors capture this demand, we offer:
✅ Competitive pricing
✅ Reliable stock & on-time shipping
✅ Region-specific product recommendations

👉 Should I prepare a *customized product list* for your country/region?

Warm regards,
Rahul Singh
Digimed Exxim
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    elif day == 40:
        subject = "Reserve your stock + 5% discount this month"
        body = """Dear Sir/Mam,

We’re closing this month’s export shipping schedule and wanted to check if you’d like me to reserve a slot for you.

To make this more valuable, we’re offering an *exclusive 5% discount on your first confirmed order this month.*

Early confirmation gives you:
✅ Priority shipping
✅ Assured stock allocation
✅ Discount on first order
✅ Best pricing for Steroids, Weight Loss, Tapentadol, Carisoprodol & Herbal

👉 Should I block a slot and apply your discount?

Warm regards,
Rahul Singh
Digimed Exxim
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    elif day == 90:
        subject = "Should I keep you updated with new launches?"
        body = """Dear Sir/Mam,

If now isn’t the right time, that’s absolutely fine — I can keep you updated on:

* New launches in Weight Loss & Herbal ranges
* Market-specific opportunities
* Special offers for distributors

👉 Would you like me to keep sending updates?

Thank you for considering us, and I look forward to collaborating in the future.

Warm regards,
Rahul Singh
Digimed Exxim
📞 +91 9158001207 (WhatsApp/Telegram/Signal available)"""

    return subject, body

# -----------------------
# 3. Send Email
# -----------------------
def send_email(to_addresses, subject, body):
    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = ", ".join(to_addresses)   # 👈 All in "To"
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        recipients = to_addresses  # Only To, no CC

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, recipients, msg.as_string())
            print(f"📤 Email sent to {', '.join(to_addresses)} | Subject: {subject}")
            return True
    except Exception as e:
        print(f"❌ Failed to send to {', '.join(to_addresses)}: {e}")
        return False

# -----------------------
# 4. Main Logic — Strict Sequential Email Schedule
# -----------------------
def run_email_schedule():
    if not os.path.exists(EXCEL_FILE):
        print("❌ Excel file not found.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    required_columns = [
        "Saved Date & Time", "To", "Email 1 Sent", "Email 2 Sent",
        "Email 3 Sent", "Email 4 Sent", "Email 5 Sent", "Email 6 Sent"
    ]

    for col in required_columns:
        if col not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col)
            headers.append(col)

    email_schedule = [
        (1, "Email 1 Sent", 1, 1),
        (2, "Email 2 Sent", 4, 4),
        (3, "Email 3 Sent", 10, 10),
        (4, "Email 4 Sent", 20, 20),
        (5, "Email 5 Sent", 40, 40),
        (6, "Email 6 Sent", 90, 90)
    ]

    email_pattern = re.compile(r"[^@]+@[^@]+\.[^@]+")
    emails_sent_today = 0

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if emails_sent_today >= MAX_EMAILS_PER_DAY:
            print(f"⏹ Daily limit of {MAX_EMAILS_PER_DAY} emails reached.")
            break

        try:
            to_field = row[headers.index("To")].value
            date_value = row[headers.index("Saved Date & Time")].value

            if not to_field:
                continue

            emails = [e.strip() for e in str(to_field).split(",") if e.strip()]
            valid_emails = [e for e in emails if email_pattern.match(e)]

            if not valid_emails:
                continue

            to_emails = valid_emails

            if isinstance(date_value, datetime):
                date_added = date_value
            else:
                try:
                    date_added = datetime.strptime(str(date_value), "%Y-%m-%d %H:%M:%S")
                except:
                    date_added = datetime.strptime(str(date_value), "%d-%m-%Y %H:%M")

            now = datetime.now()
            days_since = (now - date_added).days

            for idx, col_name, threshold, template_day in email_schedule:
                col_index = headers.index(col_name)
                sent_status = row[col_index].value

                # ✅ STRICT rule: only send if previous email was sent
                if idx > 1:
                    prev_status = row[headers.index(f"Email {idx-1} Sent")].value
                    if not prev_status or str(prev_status).strip().lower() != "sent":
                        break  # stop, because previous wasn't sent

                if sent_status and str(sent_status).strip().lower() == "sent":
                    continue  # already sent

                if days_since >= threshold:
                    subject, body = get_email_template(template_day)
                    success = send_email(to_emails, subject, body)
                    if success:
                        ws.cell(row=i, column=col_index + 1, value="Sent")
                        emails_sent_today += 1

                        # 🌟 Randomized Delay (12–18 mins between each email)
                        if emails_sent_today < MAX_EMAILS_PER_DAY:
                            delay = random.randint(720, 1080)
                            print(f"⏳ Waiting {delay//60} minutes before next email...")
                            time.sleep(delay)
                    break

        except Exception as e:
            print(f"❌ Error processing row {i}: {e}")
            continue

    wb.save(EXCEL_FILE)
    print(f"✅ Email automation run complete. {emails_sent_today} emails sent today.")

# -----------------------
# 5. Trigger
# -----------------------
if __name__ == "__main__":
    run_email_schedule()

