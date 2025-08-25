import email
import imaplib
import re
import openpyxl
import os
from datetime import datetime

# ----------------------------
# 1. Credentials and Settings
# ----------------------------
EMAIL = "digimedexxim@gmail.com"
PASSWORD = "bbdt pwfs hqok wflp"
IMAP_SERVER = "imap.gmail.com"
SEARCH_KEYWORD = '(FROM "buyleads@indiamart.com")'
OUTPUT_XLSX = "indiamart_leads.xlsx"

# ----------------------------
# 2. Excel Initialization
# ----------------------------
HEADERS = [
    "Date", "Name", "Phone","WhatsApp 1 Sent", "WhatsApp 2 Sent", "WhatsApp 3 Sent", "WhatsApp 4 Sent", "WhatsApp 5 Sent"
]

if not os.path.exists(OUTPUT_XLSX):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    wb.save(OUTPUT_XLSX)

# ----------------------------
# 3. Utility Functions
# ----------------------------
def clean_text(text):
    return re.sub(r'[\r\n]+', '\n', text).strip()

def filter_irrelevant_lines(body):
    skip_patterns = [
        r"Phone ✓ Email ✓",
        r"Email ✓",
        r"Member Since: \d+ (month|months)",
        r"\|? ?Reply To This Message",
        r"IndiaMART recommends.*",
        r"If you would like to unsubscribe.*",
        r"https://help\.indiamart\.com",
        r"Visit : .*",
        r"IndiaMART InterMESH Ltd\.",
        r"Call Us :.*",
        r"Email: buyleads@indiamart.com"
    ]
    lines = body.split('\n')
    return '\n'.join(line for line in lines if not any(re.search(pat, line.strip()) for pat in skip_patterns))

def extract_lead_info(body):
    body = filter_irrelevant_lines(body)
    lines = body.split('\n')

    name = phone = email_addr = ""

    for line in lines:
        line = line.strip()

        # Get name
        if not name and re.match(r'^[A-Za-z ]{3,}$', line):
            name = line

        # Extract phone
        if "Click to call:" in line and not phone:
            phone = line.split("Click to call:")[1].strip()

        # Extract email
        if not email_addr:
            match = re.search(r'[\w\.-]+@[\w\.-]+', line)
            if match:
                email_addr = match.group().strip()

        if name and phone and email_addr:
            break

    return name, phone, email_addr

def deduplicate_rows(rows):
    seen = set()
    deduped = []
    for row in reversed(rows):
        if row and len(row) >= 4 and row[1] and row[3]:
            key = (row[1].strip(), row[3].strip().lower())
            if key not in seen:
                seen.add(key)
                deduped.insert(0, row)
    return deduped

# ----------------------------
# 4. Fetch Leads with Pagination
# ----------------------------
def fetch_leads():
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL, PASSWORD)
    mail.select("inbox")

    status, messages = mail.search(None, SEARCH_KEYWORD)
    if status != "OK":
        print("❌ No messages found!")
        return

    mail_ids = messages[0].split()
    total_found = len(mail_ids)
    print(f"📥 Total matching leads found: {total_found}")

    wb = openpyxl.load_workbook(OUTPUT_XLSX)
    ws = wb.active
    existing_rows = list(ws.iter_rows(min_row=2, values_only=True))

    existing_keys = set(
        (r[1].strip(), r[3].strip().lower())
        for r in existing_rows
        if r and len(r) >= 4 and r[1] and r[3]
    )

    new_rows = []
    BATCH_SIZE = 50  # process in chunks to avoid Gmail timeout

    for start in range(0, total_found, BATCH_SIZE):
        batch_ids = mail_ids[start:start + BATCH_SIZE]
        for num in reversed(batch_ids):
            status, msg_data = mail.fetch(num, '(RFC822)')
            if status != "OK":
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            body = ""

            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain" and "attachment" not in str(part.get("Content-Disposition")):
                        try:
                            body = part.get_payload(decode=True).decode(errors="ignore")
                            break
                        except:
                            continue
            else:
                try:
                    body = msg.get_payload(decode=True).decode(errors="ignore")
                except:
                    continue

            body = clean_text(body)
            name, phone, email_addr = extract_lead_info(body)
            date_str = datetime.now().strftime("%d-%m-%Y %H:%M")

            key = (name.strip(), email_addr.strip().lower())
            if name and (phone or email_addr) and key not in existing_keys:
                new_rows.append([date_str, name, phone, email_addr])
                existing_keys.add(key)

        print(f"✅ Processed {min(start + BATCH_SIZE, total_found)} / {total_found} leads so far...")

    all_rows = existing_rows + new_rows
    deduped_rows = deduplicate_rows(all_rows)

    # Save to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in deduped_rows:
        ws.append(row)
    wb.save(OUTPUT_XLSX)

    mail.logout()
    print(f"🎯 {len(new_rows)} new leads saved. Total entries after deduplication: {len(deduped_rows)}")

# ----------------------------
# 5. Run Script
# ----------------------------
if __name__ == "__main__":
    fetch_leads()
