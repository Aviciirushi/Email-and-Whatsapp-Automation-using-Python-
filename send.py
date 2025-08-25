import imaplib
import email
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# ----------------------------
# 1. Gmail Credentials
# ----------------------------
# 👉 Replace with your own Gmail + App Password
EMAIL_USER = "your_email@gmail.com"
EMAIL_PASS = "your_app_password"

# ----------------------------
# 2. Excel File Setup
# ----------------------------
EXCEL_FILE = "sent_emails_list.xlsx"

if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # Ensure header exists
    if ws.max_row == 0 or ws.cell(row=1, column=1).value != "Saved Date & Time":
        ws.append([
            "Saved Date & Time", "To",
            "Email 1 Sent", "Email 2 Sent", "Email 3 Sent",
            "Email 4 Sent", "Email 5 Sent"
        ])
    # Load existing recipients (normalized for uniqueness)
    saved_recipients = {
        str(row[1].value).strip().lower()
        for row in ws.iter_rows(min_row=2) if row[1].value
    }
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sent Emails"
    ws.append([
        "Saved Date & Time", "To",
        "Email 1 Sent", "Email 2 Sent", "Email 3 Sent",
        "Email 4 Sent", "Email 5 Sent"
    ])
    saved_recipients = set()

# ----------------------------
# 3. Connect to Gmail (IMAP)
# ----------------------------
mail = imaplib.IMAP4_SSL("imap.gmail.com")
mail.login(EMAIL_USER, EMAIL_PASS)
mail.select('"[Gmail]/Sent Mail"')  # Access Gmail's Sent Mail folder

# Fetch all sent emails
status, messages = mail.search(None, "ALL")
email_ids = messages[0].split()

print(f"📧 Found {len(email_ids)} sent emails in Gmail.")

saved_count = 0

# ----------------------------
# 4. Process Sent Emails
# ----------------------------
for eid in email_ids:
    status, msg_data = mail.fetch(eid, "(RFC822)")
    for response_part in msg_data:
        if isinstance(response_part, tuple):
            msg = email.message_from_bytes(response_part[1])

            # Get recipients (raw format from Gmail)
            to_ = msg.get("To", "Unknown").strip()

            # Normalize for uniqueness
            to_normalized = to_.lower().strip()

            # Save only if not already recorded
            if to_normalized not in saved_recipients:
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([current_datetime, to_, "", "", "", "", ""])
                saved_recipients.add(to_normalized)
                saved_count += 1
                print(f"✅ Saved new entry: {to_}")

# ----------------------------
# 5. Save & Logout
# ----------------------------
wb.save(EXCEL_FILE)
print(f"💾 Saved {saved_count} new unique recipients to {EXCEL_FILE}")

mail.logout()
